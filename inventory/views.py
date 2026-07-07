from decimal import Decimal
from io import BytesIO
import qrcode
from PIL import Image, ImageDraw, ImageFont
from xml.sax.saxutils import escape
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from django.http import HttpResponse
from rest_framework.decorators import action
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    TransportOrder,
    PackagingType,
    StorageStrategySettings,
    Product,
    InventoryTransaction,
    StockMovement,
    InventorySession,
    InventoryCount,
    StorageLocation,
    StorageLocationStock,
    Supplier,
    PurchaseOrder,
    PurchaseOrderItem,
    Customer,
    CustomerContact,
    DeliveryAddress,
    CustomerNote,
    AuditLog,
)
from .serializers import (
    TransportOrderSerializer,
    PackagingTypeSerializer,
    StorageStrategySettingsSerializer,
    ProductSerializer,
    InventoryTransactionSerializer,
    StockMovementSerializer,
    InventorySessionSerializer,
    InventoryCountSerializer,
    StorageLocationSerializer,
    StorageLocationStockSerializer,
    SupplierSerializer,
    PurchaseOrderSerializer,
    PurchaseOrderItemSerializer,

    CustomerSerializer,
    CustomerContactSerializer,
    DeliveryAddressSerializer,
    CustomerNoteSerializer,
    AdminUserSerializer,
    AuditLogSerializer,
)
from .permissions import IsAdmin, IsLagerOrAdmin, IsEinkaufOrAdmin


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        data = super().validate(attrs)

        user = self.user
        profile = getattr(user, "userprofile", None)

        data["user"] = {
            "username": user.username,
            "role": profile.role if profile else "viewer",
        }

        return data



class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


def build_qr_png_response(payload: str, filename: str, label_lines: list[str] | None = None) -> HttpResponse:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(payload)
    qr.make(fit=True)

    qr_image = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    clean_label_lines = [
        line.strip()
        for line in (label_lines or [])
        if line and line.strip()
    ]

    if clean_label_lines:
        font = ImageFont.load_default()
        line_spacing = 8
        padding_x = 20
        padding_y = 18

        measure_image = Image.new("RGB", (1, 1), "white")
        measure_draw = ImageDraw.Draw(measure_image)

        text_sizes = []
        for line in clean_label_lines:
            bbox = measure_draw.textbbox((0, 0), line, font=font)
            text_sizes.append((bbox[2] - bbox[0], bbox[3] - bbox[1]))

        label_height = (
            padding_y * 2
            + sum(height for _, height in text_sizes)
            + line_spacing * max(0, len(clean_label_lines) - 1)
        )

        final_width = max(qr_image.width, max(width for width, _ in text_sizes) + padding_x * 2)
        final_height = qr_image.height + label_height

        final_image = Image.new("RGB", (final_width, final_height), "white")
        final_image.paste(qr_image, ((final_width - qr_image.width) // 2, 0))

        draw = ImageDraw.Draw(final_image)
        y = qr_image.height + padding_y

        for line, (text_width, text_height) in zip(clean_label_lines, text_sizes):
            x = (final_width - text_width) // 2
            draw.text((x, y), line, fill="black", font=font)
            y += text_height + line_spacing

        image = final_image
    else:
        image = qr_image

    buffer = BytesIO()
    image.save(buffer, format="PNG")

    response = HttpResponse(buffer.getvalue(), content_type="image/png")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    response["X-QR-Payload"] = payload
    return response

class StorageLocationViewSet(viewsets.ModelViewSet):
    queryset = StorageLocation.objects.all().order_by("code")
    serializer_class = StorageLocationSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsEinkaufOrAdmin()]

    @action(detail=True, methods=["get"], url_path="qr-code")
    def qr_code(self, request, pk=None):
        location = self.get_object()

        payload = f"LOCATION:{location.id}|CODE:{location.code}"
        if location.name:
            payload += f"|NAME:{location.name}"

        safe_code = slugify(location.code or f"location-{location.id}") or f"location-{location.id}"
        filename = f"storage-location-{safe_code}-qr.png"

        return build_qr_png_response(
            payload,
            filename,
            label_lines=[
                f"Lagerort: {location.code}",
                location.name or "",
            ],
        )

class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all().order_by("name")
    serializer_class = SupplierSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsEinkaufOrAdmin()]




class PurchaseOrderViewSet(viewsets.ModelViewSet):
    queryset = (
        PurchaseOrder.objects.select_related(
            "supplier",
            "created_by",
            "released_by",
            "ordered_by",
            "received_by",
        )
        .prefetch_related("items__product")
        .order_by("-created_at", "-id")
    )
    serializer_class = PurchaseOrderSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsEinkaufOrAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"], url_path="release")
    def release(self, request, pk=None):
        order = self.get_object()

        if order.status != "DRAFT":
            return Response(
                {"detail": "Nur Entwürfe können freigegeben werden."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order.status = "RELEASED"
        order.released_by = request.user
        order.released_at = timezone.now()
        order.save(
            update_fields=[
                "status",
                "released_by",
                "released_at",
                "updated_at",
            ]
        )

        return Response(self.get_serializer(order).data)

    @action(detail=True, methods=["post"], url_path="mark-ordered")
    def mark_ordered(self, request, pk=None):
        order = self.get_object()

        if order.status not in ["DRAFT", "RELEASED"]:
            return Response(
                {"detail": "Diese Bestellung kann nicht auf Bestellt gesetzt werden."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order.status = "ORDERED"
        order.ordered_by = request.user
        order.ordered_at = timezone.now()
        order.save(
            update_fields=[
                "status",
                "ordered_by",
                "ordered_at",
                "updated_at",
            ]
        )

        return Response(self.get_serializer(order).data)

    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel(self, request, pk=None):
        order = self.get_object()

        if order.status in ["RECEIVED", "CANCELLED"]:
            return Response(
                {"detail": "Gelieferte oder stornierte Bestellungen können nicht storniert werden."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order.status = "CANCELLED"
        order.save(update_fields=["status", "updated_at"])

        return Response(self.get_serializer(order).data)


class PurchaseOrderItemViewSet(viewsets.ModelViewSet):
    queryset = (
        PurchaseOrderItem.objects.select_related(
            "purchase_order",
            "product",
        )
        .order_by("purchase_order__order_number", "product__name")
    )
    serializer_class = PurchaseOrderItemSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsEinkaufOrAdmin()]

    @action(detail=True, methods=["post"], url_path="receive")
    @transaction.atomic
    def receive(self, request, pk=None):
        item = self.get_object()
        order = item.purchase_order

        if order.status == "CANCELLED":
            return Response(
                {"detail": "Stornierte Bestellungen können nicht empfangen werden."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if order.status == "RECEIVED":
            return Response(
                {"detail": "Diese Bestellung ist bereits vollständig geliefert."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        open_quantity = item.open_quantity

        try:
            quantity = int(request.data.get("quantity") or open_quantity)
        except (TypeError, ValueError):
            return Response(
                {"quantity": "Die Empfangsmenge muss eine ganze Zahl sein."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if quantity <= 0:
            return Response(
                {"quantity": "Die Empfangsmenge muss größer als 0 sein."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if quantity > open_quantity:
            return Response(
                {
                    "quantity": (
                        f"Es können maximal {open_quantity} Stück aus dieser "
                        "Bestellposition empfangen werden."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        unit_purchase_price = request.data.get("unit_purchase_price")
        if unit_purchase_price in ["", None] and item.unit_price is not None:
            unit_purchase_price = str(item.unit_price)

        reference_number = (
            request.data.get("reference_number")
            or f"WE-{order.order_number or order.id}"
        )

        note = (
            request.data.get("note")
            or (
                f"Wareneingang aus Bestellung {order.order_number or order.id}, "
                f"Position #{item.id}."
            )
        )

        movement_serializer = StockMovementSerializer(
            data={
                "product": item.product_id,
                "movement_type": "IN",
                "quantity": quantity,
                "storage_location": request.data.get("storage_location"),
                "packaging_type": request.data.get("packaging_type"),
                "load_carrier_type": request.data.get("load_carrier_type"),
                "packaging_quantity": request.data.get("packaging_quantity") or 1,
                "unit_purchase_price": unit_purchase_price,
                "expiry_date": request.data.get("expiry_date") or None,
                "reference_number": reference_number,
                "note": note,
            }
        )
        movement_serializer.is_valid(raise_exception=True)
        movement = movement_serializer.save(created_by=request.user)

        item.received_quantity += quantity
        item.save(update_fields=["received_quantity", "updated_at"])

        order_items = list(order.items.all())
        total_open_quantity = sum(order_item.open_quantity for order_item in order_items)
        total_received_quantity = sum(
            order_item.received_quantity for order_item in order_items
        )

        update_fields = ["status", "received_by", "received_at", "updated_at"]

        if total_open_quantity == 0:
            order.status = "RECEIVED"
        elif total_received_quantity > 0:
            order.status = "PARTIALLY_RECEIVED"

        order.received_by = request.user
        order.received_at = timezone.now()
        order.save(update_fields=update_fields)

        return Response(
            {
                "movement": StockMovementSerializer(movement).data,
                "item": self.get_serializer(item).data,
                "order_status": order.status,
            },
            status=status.HTTP_201_CREATED,
        )



class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all().order_by("name")
    serializer_class = CustomerSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsEinkaufOrAdmin()]


class CustomerContactViewSet(viewsets.ModelViewSet):
    queryset = CustomerContact.objects.select_related("customer").order_by(
        "customer__name",
        "last_name",
        "first_name",
    )
    serializer_class = CustomerContactSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsEinkaufOrAdmin()]


class DeliveryAddressViewSet(viewsets.ModelViewSet):
    queryset = DeliveryAddress.objects.select_related("customer").order_by(
        "customer__name",
        "-is_default",
        "label",
    )
    serializer_class = DeliveryAddressSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsEinkaufOrAdmin()]


class CustomerNoteViewSet(viewsets.ModelViewSet):
    queryset = CustomerNote.objects.select_related(
        "customer",
        "created_by",
    ).order_by("-created_at")
    serializer_class = CustomerNoteSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsEinkaufOrAdmin()]

def perform_create(self, serializer):
    movement = serializer.save(created_by=self.request.user)

    product = movement.product

    # Nur bei Wareneingang
    if movement.movement_type == "IN":

        suggested_location = self.get_storage_location_suggestion(product)

        if suggested_location:

            # Produkt Lagerplatz setzen
            product.storage_location = suggested_location

            # Lagerplatz als belegt markieren
            suggested_location.is_empty = False
            suggested_location.save(update_fields=["is_empty"])

            product.save(update_fields=["storage_location"])

        else:
            raise ValueError(
                "Kein geeigneter Lagerplatz gefunden."
            )

class AdminUserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.select_related("userprofile").order_by("username")
    serializer_class = AdminUserSerializer

    def get_permissions(self):
        return [IsAuthenticated(), IsAdmin()]

    def perform_create(self, serializer):
        created_user = serializer.save()

        AuditLog.objects.create(
            area="Benutzerverwaltung",
            action="CREATE",
            object_type="User",
            object_id=str(created_user.id),
            message=f"Benutzer {created_user.username} wurde angelegt.",
            created_by=self.request.user,
            metadata={"username": created_user.username},
        )

    def perform_update(self, serializer):
        updated_user = serializer.save()

        AuditLog.objects.create(
            area="Benutzerverwaltung",
            action="UPDATE",
            object_type="User",
            object_id=str(updated_user.id),
            message=f"Benutzer {updated_user.username} wurde geändert.",
            created_by=self.request.user,
            metadata={"username": updated_user.username},
        )


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.select_related("created_by").order_by("-created_at")
    serializer_class = AuditLogSerializer

    def get_permissions(self):
        return [IsAuthenticated(), IsAdmin()]


class TransportOrderViewSet(viewsets.ModelViewSet):
    queryset = (
        TransportOrder.objects.select_related(
            "product",
            "source_location",
            "target_location",
            "assigned_to",
            "created_by",
        )
        .all()
    )
    serializer_class = TransportOrderSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def _parse_scan_value(self, raw_value: str) -> dict[str, str]:
        scan_parts: dict[str, str] = {}

        for part in raw_value.split("|"):
            if ":" not in part:
                continue

            key, value = part.split(":", 1)
            key = key.strip().upper()
            value = value.strip()

            if key and value:
                scan_parts[key] = value

        return scan_parts

    def _find_product_from_scan(self, raw_value: str):
        scan_parts = self._parse_scan_value(raw_value)
        is_structured_scan = "|" in raw_value or ":" in raw_value

        product_id = scan_parts.get("PRODUCT")
        sku = scan_parts.get("SKU") or ("" if is_structured_scan else raw_value)

        products = Product.objects.all()

        if product_id:
            product = products.filter(id=product_id).first()
            if product:
                return product

        if sku:
            product = products.filter(sku__iexact=sku).first()
            if product:
                return product

        return None

    def _find_location_from_scan(self, raw_value: str):
        scan_parts = self._parse_scan_value(raw_value)
        is_structured_scan = "|" in raw_value or ":" in raw_value

        location_id = scan_parts.get("LOCATION")
        code = scan_parts.get("CODE") or ("" if is_structured_scan else raw_value)

        locations = StorageLocation.objects.all()

        if location_id:
            location = locations.filter(id=location_id).first()
            if location:
                return location

        if code:
            location = locations.filter(code__iexact=code).first()
            if location:
                return location

        return None

    def _error_response(self, order, message: str, scan_value: str, status_code=status.HTTP_400_BAD_REQUEST):
        order.last_scan_value = scan_value
        order.last_error = message
        order.save(update_fields=["last_scan_value", "last_error", "updated_at"])

        return Response(
            {
                "detail": message,
                "status": order.status,
                "expected_source_location": order.source_location.code,
                "expected_target_location": order.target_location.code if order.target_location else None,
                "last_scan_value": scan_value,
                "warning": "BEEP",
            },
            status=status_code,
        )

    def _get_source_stock_for_outbound(self, product, quantity):
        queryset = (
            StorageLocationStock.objects.select_related("storage_location")
            .filter(
                product=product,
                quantity__gte=quantity,
                storage_location__is_active=True,
                storage_location__is_blocked=False,
            )
        )

        strategy = getattr(product, "removal_strategy", "FIFO")

        if strategy == "LIFO":
            queryset = queryset.order_by("-created_at")
        elif strategy == "FEFO":
            queryset = queryset.order_by("expiry_date", "created_at")
        elif strategy == "HIFO":
            queryset = queryset.order_by("-unit_purchase_price", "created_at")
        elif strategy == "LOFO":
            queryset = queryset.order_by("unit_purchase_price", "created_at")
        else:
            queryset = queryset.order_by("created_at")

        return queryset.first()

    @action(detail=False, methods=["post"], url_path="create-from-outbound")
    @transaction.atomic
    def create_from_outbound(self, request):
        product_id = request.data.get("product")
        quantity_value = request.data.get("quantity")
        target_location_id = request.data.get("target_location")
        reference_number = request.data.get("reference_number", "")

        if not product_id:
            return Response(
                {"detail": "Bitte ein Produkt angeben."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            quantity = Decimal(str(quantity_value))
        except Exception:
            return Response(
                {"detail": "Bitte eine gültige Menge angeben."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if quantity <= 0:
            return Response(
                {"detail": "Die Menge muss größer als 0 sein."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        product = Product.objects.filter(id=product_id).first()

        if not product:
            return Response(
                {"detail": "Produkt wurde nicht gefunden."},
                status=status.HTTP_404_NOT_FOUND,
            )

        source_stock = self._get_source_stock_for_outbound(product, quantity)

        if not source_stock:
            return Response(
                {
                    "detail": "Kein geeigneter Entnahmeplatz mit ausreichendem Bestand gefunden.",
                    "product": product.name,
                    "quantity": str(quantity),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        target_location = None

        if target_location_id:
            target_location = StorageLocation.objects.filter(id=target_location_id).first()

            if not target_location:
                return Response(
                    {"detail": "Zielplatz wurde nicht gefunden."},
                    status=status.HTTP_404_NOT_FOUND,
                )

        order = TransportOrder.objects.create(
            product=product,
            quantity=quantity,
            source_location=source_stock.storage_location,
            target_location=target_location,
            reference_number=reference_number,
            created_by=request.user,
            status=TransportOrder.Status.CREATED,
        )

        serializer = self.get_serializer(order)

        return Response(
            {
                "detail": "Transportauftrag wurde automatisch erstellt.",
                "transport_order": serializer.data,
                "next_step": "Stapler-Terminal öffnen und Quellplatz scannen.",
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=["get"], url_path="active")
    def active(self, request):
        queryset = (
            self.get_queryset()
            .exclude(
                status__in=[
                    TransportOrder.Status.COMPLETED,
                    TransportOrder.Status.CANCELLED,
                ]
            )
            .order_by("priority", "created_at")
        )

        serializer = self.get_serializer(queryset, many=True)

        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="assign-to-me")
    def assign_to_me(self, request, pk=None):
        order = self.get_object()

        if order.status in [
            TransportOrder.Status.COMPLETED,
            TransportOrder.Status.CANCELLED,
        ]:
            return Response(
                {"detail": "Abgeschlossene oder stornierte Transportaufträge können nicht übernommen werden."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order.assigned_to = request.user

        if order.status == TransportOrder.Status.CREATED:
            order.status = TransportOrder.Status.ASSIGNED

        order.save(update_fields=["assigned_to", "status", "updated_at"])

        serializer = self.get_serializer(order)

        return Response(
            {
                "detail": "Transportauftrag wurde dem aktuellen Stapler-Benutzer zugewiesen.",
                "transport_order": serializer.data,
                "next_step": "Quellplatz scannen.",
            }
        )

    @action(detail=True, methods=["post"], url_path="scan")
    @transaction.atomic
    def scan(self, request, pk=None):
        order = self.get_object()
        scan_value = str(request.data.get("scan_value", "")).strip()

        if not scan_value:
            return self._error_response(
                order,
                "Bitte einen Barcode oder QR-Code scannen.",
                scan_value,
            )

        if order.status in [
            TransportOrder.Status.COMPLETED,
            TransportOrder.Status.CANCELLED,
        ]:
            return self._error_response(
                order,
                "Dieser Transportauftrag ist bereits abgeschlossen oder storniert.",
                scan_value,
            )

        scanned_location = self._find_location_from_scan(scan_value)
        scanned_product = self._find_product_from_scan(scan_value)

        if order.status in [
            TransportOrder.Status.CREATED,
            TransportOrder.Status.ASSIGNED,
            TransportOrder.Status.ERROR,
        ]:
            if scanned_location and scanned_location.id == order.source_location_id:
                order.status = TransportOrder.Status.IN_TRANSIT
                order.assigned_to = order.assigned_to or request.user
                order.picked_at = timezone.now()
                order.last_scan_value = scan_value
                order.last_error = ""
                order.save(
                    update_fields=[
                        "status",
                        "assigned_to",
                        "picked_at",
                        "last_scan_value",
                        "last_error",
                        "updated_at",
                    ]
                )

                serializer = self.get_serializer(order)

                return Response(
                    {
                        "detail": "Quellplatz bestätigt. Ware wurde aufgenommen.",
                        "transport_order": serializer.data,
                        "next_step": "Zum Zielplatz fahren und Zielplatz scannen.",
                    }
                )

            if scanned_location:
                return self._error_response(
                    order,
                    f"Falscher Quellplatz. Erwartet: {order.source_location.code}, gescannt: {scanned_location.code}.",
                    scan_value,
                )

            if scanned_product:
                return self._error_response(
                    order,
                    f"Produkt erkannt: {scanned_product.name}. Bitte den Quellplatz {order.source_location.code} scannen.",
                    scan_value,
                )

            return self._error_response(
                order,
                f"Scan nicht erkannt. Erwartet wird der Quellplatz {order.source_location.code}.",
                scan_value,
            )

        if order.status == TransportOrder.Status.IN_TRANSIT:
            if not order.target_location:
                return self._error_response(
                    order,
                    "Für diesen Transportauftrag ist noch kein Zielplatz hinterlegt.",
                    scan_value,
                )

            if scanned_location and scanned_location.id == order.target_location_id:
                order.status = TransportOrder.Status.COMPLETED
                order.completed_at = timezone.now()
                order.last_scan_value = scan_value
                order.last_error = ""
                order.save(
                    update_fields=[
                        "status",
                        "completed_at",
                        "last_scan_value",
                        "last_error",
                        "updated_at",
                    ]
                )

                serializer = self.get_serializer(order)

                return Response(
                    {
                        "detail": "Zielplatz bestätigt. Transportauftrag wurde abgeschlossen.",
                        "transport_order": serializer.data,
                        "next_step": "Transport abgeschlossen.",
                    }
                )

            if scanned_location:
                return self._error_response(
                    order,
                    f"Falscher Zielplatz. Erwartet: {order.target_location.code}, gescannt: {scanned_location.code}.",
                    scan_value,
                )

            return self._error_response(
                order,
                f"Scan nicht erkannt. Erwartet wird der Zielplatz {order.target_location.code}.",
                scan_value,
            )

        return self._error_response(
            order,
            f"Für den aktuellen Status {order.status} ist kein Scan-Schritt definiert.",
            scan_value,
        )

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.select_related("storage_location").order_by("-id")
    serializer_class = ProductSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsLagerOrAdmin()]

    @action(detail=True, methods=["get"], url_path="qr-code")
    def qr_code(self, request, pk=None):
        product = self.get_object()

        payload = f"PRODUCT:{product.id}|SKU:{product.sku}"
        if product.name:
            payload += f"|NAME:{product.name}"

        safe_sku = slugify(product.sku or product.name or f"product-{product.id}") or f"product-{product.id}"
        filename = f"product-{safe_sku}-qr.png"

        return build_qr_png_response(payload, filename)

class InventoryTransactionViewSet(viewsets.ModelViewSet):
    queryset = InventoryTransaction.objects.all().order_by("-id")
    serializer_class = InventoryTransactionSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsLagerOrAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class StockMovementViewSet(viewsets.ModelViewSet):
    queryset = StockMovement.objects.select_related(
        "product",
        "created_by",
        "storage_location",
        "packaging_type",
        "load_carrier_type",
    ).order_by("-created_at")

    serializer_class = StockMovementSerializer

    def get_storage_location_suggestion(self, product):
        """
        Ermittelt einen intelligenten Lagerplatzvorschlag.
        """

        # 1. Festplatzstrategie
        if (
            product.putaway_strategy == "FIXED_BIN"
            and product.fixed_storage_location
            and product.fixed_storage_location.is_active
            and not product.fixed_storage_location.is_blocked
        ):
            return product.fixed_storage_location

        # 2. Leerplatzsuche
        empty_location = (
            StorageLocation.objects.filter(
                is_active=True,
                is_blocked=False,
                is_empty=True,
            )
            .order_by("zone", "aisle", "rack", "shelf")
            .first()
        )

        if empty_location:
            return empty_location

        # 3. Zulagerung / Mischlager
        if product.putaway_strategy == "ADD_TO_STOCK":
            mixed_location = (
                StorageLocation.objects.filter(
                    is_active=True,
                    is_blocked=False,
                    allow_mixed_products=True,
                ).first()
            )

            if mixed_location:
                return mixed_location

        return None

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsLagerOrAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Bewegungshistorie"

        headers = [
            "Datum",
            "Produkt",
            "Typ",
            "Menge",
            "Referenz",
            "Lagerplatz",
            "Verpackung",
            "Ladungsträger",
            "Packmenge",
            "Kosten",
            "Notiz",
            "Benutzer",
        ]

        worksheet.append(headers)

        column_widths = {
            "A": 22,
            "B": 28,
            "C": 18,
            "D": 12,
            "E": 18,
            "F": 45,
            "G": 18,
            "H": 22,
            "I": 14,
            "J": 14,
            "K": 45,
            "L": 18,
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1E293B",
        )
        header_font = Font(
            bold=True,
            color="FFFFFF",
        )
        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border

        movements = self.get_queryset()

        for movement in movements:
            movement_type = (
                "Wareneingang"
                if movement.movement_type == "IN"
                else "Warenausgang"
            )

            quantity = (
                -movement.quantity
                if movement.movement_type == "OUT"
                else movement.quantity
            )

            row = [
                movement.created_at.strftime("%d.%m.%Y %H:%M:%S"),
                movement.product.name if movement.product else "",
                movement_type,
                quantity,
                movement.reference_number or "",
                str(movement.storage_location) if movement.storage_location else "",
                movement.packaging_type.name if movement.packaging_type else "",
                movement.load_carrier_type.name if movement.load_carrier_type else "",
                movement.packaging_quantity
                if movement.packaging_type or movement.load_carrier_type
                else "",
                movement.packaging_cost_total
                if movement.packaging_cost_total is not None
                else "",
                movement.note or "",
                movement.created_by.username if movement.created_by else "",
            ]

            worksheet.append(row)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        for cell in worksheet["J"][1:]:
            cell.number_format = '#,##0.00 €'

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            'attachment; filename="bewegungshistorie.xlsx"'
        )

        workbook.save(response)
        return response

    serializer_class = StockMovementSerializer

    def get_storage_location_suggestion(self, product):
        """
        Ermittelt einen intelligenten Lagerplatzvorschlag.
        """

        # 1. Festplatzstrategie
        if (
            product.putaway_strategy == "FIXED_BIN"
            and product.fixed_storage_location
            and product.fixed_storage_location.is_active
            and not product.fixed_storage_location.is_blocked
        ):
            return product.fixed_storage_location

        # 2. Leerplatzsuche
        empty_location = (
            StorageLocation.objects.filter(
                is_active=True,
                is_blocked=False,
                is_empty=True,
            )
            .order_by("zone", "aisle", "rack", "shelf")
            .first()
        )

        if empty_location:
            return empty_location

        # 3. Zulagerung / Mischlager
        if product.putaway_strategy == "ADD_TO_STOCK":
            mixed_location = (
                StorageLocation.objects.filter(
                    is_active=True,
                    is_blocked=False,
                    allow_mixed_products=True,
                ).first()
            )

            if mixed_location:
                return mixed_location

        return None

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsLagerOrAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)



class InventorySessionViewSet(viewsets.ModelViewSet):
    queryset = (
        InventorySession.objects.select_related("created_by")
        .prefetch_related("counts")
        .order_by("-created_at")
    )
    serializer_class = InventorySessionSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsLagerOrAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"], url_path="complete")
    def complete(self, request, pk=None):
        session = self.get_object()
        session.status = "COMPLETED"
        session.completed_at = timezone.now()
        session.save(update_fields=["status", "completed_at"])

        serializer = self.get_serializer(session)
        return Response(serializer.data)


    @action(detail=True, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request, pk=None):
        session = self.get_object()

        counts = (
            session.counts.select_related(
                "product",
                "created_by",
                "correction_movement",
            )
            .order_by("product__name")
        )

        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=24,
            leftMargin=24,
            topMargin=24,
            bottomMargin=24,
        )

        styles = getSampleStyleSheet()
        story = []

        title = Paragraph(
            "Smart Inventory Manager - Inventurbericht",
            styles["Title"],
        )
        story.append(title)
        story.append(Spacer(1, 12))

        created_at = session.created_at
        if timezone.is_aware(created_at):
            created_at = timezone.localtime(created_at)

        completed_at = session.completed_at
        if completed_at and timezone.is_aware(completed_at):
            completed_at = timezone.localtime(completed_at)

        meta_data = [
            ["Inventur", escape(session.title or "")],
            ["Status", "Abgeschlossen" if session.status == "COMPLETED" else "Offen"],
            ["Erstellt von", escape(session.created_by.username if session.created_by else "")],
            ["Erstellt am", created_at.strftime("%d.%m.%Y %H:%M")],
            [
                "Abgeschlossen am",
                completed_at.strftime("%d.%m.%Y %H:%M") if completed_at else "",
            ],
            ["Notiz", escape(session.note or "")],
        ]

        meta_table = Table(meta_data, colWidths=[120, 560])
        meta_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DBEAFE")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        story.append(meta_table)
        story.append(Spacer(1, 18))

        rows = [
            [
                "Produkt",
                "SKU",
                "Soll",
                "Ist",
                "Differenz",
                "Einheit",
                "Status",
                "Korrigiert",
                "Notiz",
            ]
        ]

        total_positions = 0
        counted_positions = 0
        difference_positions = 0
        corrected_positions = 0

        for count in counts:
            total_positions += 1

            if count.counted_quantity is not None:
                counted_positions += 1

            difference = count.difference

            if difference not in (None, 0):
                difference_positions += 1

            if count.correction_movement_id:
                corrected_positions += 1

            rows.append(
                [
                    Paragraph(escape(count.product.name), styles["BodyText"]),
                    escape(count.product.sku or ""),
                    count.expected_quantity,
                    count.counted_quantity if count.counted_quantity is not None else "",
                    difference if difference is not None else "",
                    escape(count.product.unit or ""),
                    "OK" if difference == 0 else "Differenz",
                    "Ja" if count.correction_movement_id else "Nein",
                    Paragraph(escape(count.note or ""), styles["BodyText"]),
                ]
            )

        table = Table(
            rows,
            repeatRows=1,
            colWidths=[150, 90, 55, 55, 65, 60, 75, 75, 190],
        )

        table_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )

        for row_index in range(1, len(rows)):
            status = rows[row_index][6]

            if status == "OK":
                table_style.add(
                    "BACKGROUND",
                    (0, row_index),
                    (-1, row_index),
                    colors.HexColor("#DCFCE7"),
                )
            else:
                table_style.add(
                    "BACKGROUND",
                    (0, row_index),
                    (-1, row_index),
                    colors.HexColor("#FEF3C7"),
                )

        table.setStyle(table_style)

        story.append(table)
        story.append(Spacer(1, 18))

        summary_rows = [
            ["Zusammenfassung", ""],
            ["Positionen gesamt", total_positions],
            ["Gezählte Positionen", counted_positions],
            ["Positionen mit Differenz", difference_positions],
            ["Korrigierte Positionen", corrected_positions],
        ]

        summary_table = Table(summary_rows, colWidths=[180, 120])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        story.append(summary_table)

        doc.build(story)

        buffer.seek(0)

        filename = f"inventurbericht-{session.id}.pdf"

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        session = self.get_object()

        counts = (
            session.counts.select_related(
                "product",
                "created_by",
                "correction_movement",
            )
            .order_by("product__name")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventurbericht"

        title_fill = PatternFill("solid", fgColor="1E3A8A")
        header_fill = PatternFill("solid", fgColor="2563EB")
        green_fill = PatternFill("solid", fgColor="DCFCE7")
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        yellow_fill = PatternFill("solid", fgColor="FEF3C7")

        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)

        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        ws.merge_cells("A1:I1")
        ws["A1"] = "Smart Inventory Manager - Inventurbericht"
        ws["A1"].fill = title_fill
        ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = "Inventur:"
        ws["B3"] = session.title

        ws["A4"] = "Status:"
        ws["B4"] = session.get_status_display()

        ws["A5"] = "Erstellt von:"
        ws["B5"] = session.created_by.username if session.created_by else "—"

        ws["A6"] = "Erstellt am:"
        ws["B6"] = session.created_at.strftime("%d.%m.%Y %H:%M")

        if session.completed_at:
            ws["A7"] = "Abgeschlossen am:"
            ws["B7"] = session.completed_at.strftime("%d.%m.%Y %H:%M")

        for cell in ["A3", "A4", "A5", "A6", "A7"]:
            ws[cell].font = bold_font

        start_row = 9

        headers = [
            "Produkt",
            "SKU",
            "Soll-Bestand",
            "Ist-Bestand",
            "Differenz",
            "Einheit",
            "Korrigiert",
            "Notiz",
            "Gezählt von",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_index, count in enumerate(counts, start=start_row + 1):
            difference = count.difference

            row_values = [
                count.product.name,
                count.product.sku,
                count.expected_quantity,
                count.counted_quantity if count.counted_quantity is not None else "",
                difference if difference is not None else "",
                count.product.unit,
                "Ja" if count.corrected else "Nein",
                count.note,
                count.created_by.username if count.created_by else "—",
            ]

            if count.counted_quantity is None:
                row_fill = yellow_fill
            elif difference != 0:
                row_fill = red_fill
            else:
                row_fill = green_fill

            for col, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_index, column=col, value=value)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["H"].width = 35
        ws.freeze_panes = "A10"

        summary_row = start_row + counts.count() + 3

        total_counts = counts.count()
        done_counts = counts.exclude(counted_quantity__isnull=True).count()
        diff_counts = sum(
            1 for count in counts if count.difference not in (None, 0)
        )
        corrected_counts = counts.filter(corrected=True).count()

        ws.cell(
            row=summary_row,
            column=1,
            value="Zusammenfassung",
        ).font = Font(bold=True, size=14)

        summary_data = [
            ("Positionen gesamt", total_counts),
            ("Gezählte Positionen", done_counts),
            ("Positionen mit Differenz", diff_counts),
            ("Korrigierte Positionen", corrected_counts),
        ]

        for index, (label, value) in enumerate(summary_data, start=summary_row + 1):
            ws.cell(row=index, column=1, value=label).font = bold_font
            ws.cell(row=index, column=2, value=value)

        correction_sheet = wb.create_sheet("Korrekturen")

        correction_headers = [
            "Produkt",
            "Bewegungstyp",
            "Menge",
            "Referenz",
            "Notiz",
            "Gebucht von",
            "Datum",
        ]

        for col, header in enumerate(correction_headers, start=1):
            cell = correction_sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = thin_border

        correction_row = 2

        for count in counts:
            movement = count.correction_movement
            if not movement:
                continue

            values = [
                movement.product.name,
                movement.get_movement_type_display(),
                movement.quantity,
                movement.reference_number,
                movement.note,
                movement.created_by.username if movement.created_by else "—",
                movement.created_at.strftime("%d.%m.%Y %H:%M"),
            ]

            for col, value in enumerate(values, start=1):
                cell = correction_sheet.cell(
                    row=correction_row,
                    column=col,
                    value=value,
                )
                cell.border = thin_border

            correction_row += 1

        for col in range(1, len(correction_headers) + 1):
            correction_sheet.column_dimensions[get_column_letter(col)].width = 24

        safe_title = slugify(session.title) or "inventur"
        filename = f"inventurbericht-{session.id}-{safe_title}.xlsx"

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


class InventoryCountViewSet(viewsets.ModelViewSet):
    queryset = InventoryCount.objects.select_related(
        "session",
        "product",
        "created_by",
        "correction_movement",
    ).order_by("-created_at")
    serializer_class = InventoryCountSerializer

    def get_permissions(self):
        if self.request.method in ["GET", "HEAD", "OPTIONS"]:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsLagerOrAdmin()]

    def get_queryset(self):
        queryset = super().get_queryset()
        session_id = self.request.query_params.get("session")

        if session_id:
            queryset = queryset.filter(session_id=session_id)

        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"], url_path="apply-correction")
    @transaction.atomic
    def apply_correction(self, request, pk=None):
        inventory_count = self.get_object()

        if inventory_count.counted_quantity is None:
            return Response(
                {"detail": "Bitte zuerst eine gezählte Menge eintragen."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if inventory_count.corrected:
            return Response(
                {
                    "detail": (
                        "Für diese Inventurposition wurde bereits eine "
                        "Korrektur gebucht."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        product = inventory_count.product
        current_quantity = product.quantity
        counted_quantity = inventory_count.counted_quantity
        difference = counted_quantity - current_quantity

        if difference == 0:
            inventory_count.corrected = True
            inventory_count.save(update_fields=["corrected"])
            serializer = self.get_serializer(inventory_count)
            return Response(serializer.data)

        movement_type = "IN" if difference > 0 else "OUT"
        movement_quantity = abs(difference)

        product.quantity = counted_quantity
        product.save(update_fields=["quantity", "updated_at"])

        movement = StockMovement.objects.create(
            product=product,
            movement_type=movement_type,
            quantity=movement_quantity,
            reference_number=f"INV-{inventory_count.session_id}-{inventory_count.id}",
            note=(
                f"Inventur-Korrektur: Soll laut System "
                f"{inventory_count.expected_quantity}, "
                f"Ist gezählt {counted_quantity}, "
                f"vorheriger aktueller Bestand {current_quantity}."
            ),
            created_by=request.user,
        )

        inventory_count.corrected = True
        inventory_count.correction_movement = movement
        inventory_count.save(
            update_fields=[
                "corrected",
                "correction_movement",
                "updated_at",
            ]
        )

        serializer = self.get_serializer(inventory_count)
        return Response(serializer.data)
    


class StorageLocationStockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = (
        StorageLocationStock.objects.select_related(
            "product",
            "storage_location",
            "packaging_type",
            "load_carrier_type",
        )
        .filter(quantity__gt=0)
        .order_by("storage_location__code", "product__name")
    )
    serializer_class = StorageLocationStockSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Lagerplatzbestand"

        headers = [
            "Lagerplatz",
            "Lagerplatz-Name",
            "Zone",
            "Gang",
            "Regal",
            "Fach",
            "Status Lagerplatz",
            "Produkt",
            "SKU",
            "Menge Lagerplatz",
            "Einheit",
            "Gesamtbestand Produkt",
            "Verpackung",
            "Ladungsträger",
            "Packmenge",
            "Einstandspreis",
            "MHD",
            "Aktualisiert",
        ]

        worksheet.append(headers)

        column_widths = {
            "A": 16,
            "B": 24,
            "C": 14,
            "D": 12,
            "E": 12,
            "F": 12,
            "G": 18,
            "H": 28,
            "I": 18,
            "J": 16,
            "K": 12,
            "L": 20,
            "M": 20,
            "N": 22,
            "O": 14,
            "P": 16,
            "Q": 14,
            "R": 22,
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        header_fill = PatternFill("solid", fgColor="2563EB")
        white_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")

        for stock in self.get_queryset():
            location = stock.storage_location
            product = stock.product

            updated_at = stock.updated_at
            if timezone.is_aware(updated_at):
                updated_at = timezone.localtime(updated_at)

            worksheet.append(
                [
                    location.code,
                    location.name,
                    location.zone or "",
                    location.aisle or "",
                    location.rack or "",
                    location.shelf or "",
                    "Frei" if location.is_empty else "Belegt",
                    product.name,
                    product.sku,
                    stock.quantity,
                    product.unit,
                    product.quantity,
                    stock.packaging_type.name if stock.packaging_type else "",
                    stock.load_carrier_type.name if stock.load_carrier_type else "",
                    stock.packaging_quantity,
                    stock.unit_purchase_price if stock.unit_purchase_price is not None else "",
                    stock.expiry_date if stock.expiry_date else "",
                    updated_at.strftime("%d.%m.%Y %H:%M"),
                ]
            )

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            'attachment; filename="lagerplatzbestand.xlsx"'
        )
        workbook.save(response)
        return response


class PackagingTypeViewSet(viewsets.ModelViewSet):
    queryset = PackagingType.objects.all().order_by("name")
    serializer_class = PackagingTypeSerializer


class StorageStrategySettingsViewSet(viewsets.ModelViewSet):
    queryset = StorageStrategySettings.objects.all()
    serializer_class = StorageStrategySettingsSerializer
